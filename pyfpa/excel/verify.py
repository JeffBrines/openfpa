"""Workbook verification harness: evaluate live-formula workbook and compare to engine.

The ``formulas`` package is a dev-only dependency; it is imported lazily so
the runtime library works without it.
"""
from __future__ import annotations

import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel


class VerifyReport(BaseModel):
    passed: bool
    failures: list[str]
    max_rel_deviation: float
    lines_checked: int


def _load_formulas():
    try:
        import formulas  # noqa: PLC0415 (lazy on purpose: dev-only dependency)
    except ImportError as exc:
        raise RuntimeError(
            "workbook verification requires the 'formulas' package: pip install formulas"
        ) from exc
    return formulas


def verify_workbook(
    path: str | Path,
    expected: pd.DataFrame,
    *,
    rel_tol: float = 1e-6,
) -> VerifyReport:
    """Evaluate the workbook's formulas in Python and compare every Model-sheet
    line that matches a column of ``expected``, month by month.

    NaN or unevaluated cells are failures, never skipped.

    Key format confirmed against formulas 1.3.4:
        ``'[<filename_exact_case>]MODEL'!<COL><ROW>``
    The sheet name is stored as-uppercase in the solution dict.
    """
    formulas = _load_formulas()
    path = Path(path)
    solution = formulas.ExcelModel().loads(str(path)).finish().calculate()

    wb = load_workbook(path)
    model_ws = wb["Model"]
    labels = {
        model_ws.cell(row=r, column=1).value: r
        for r in range(2, model_ws.max_row + 1)
    }

    # The formulas library preserves the original filename casing for the book
    # part and uppercases the sheet name. Confirmed format: "'[m.xlsx]MODEL'!B3"
    book_tag = f"[{path.name}]MODEL"

    failures: list[str] = []
    max_dev = 0.0
    lines = 0

    for line in expected.columns:
        if line not in labels:
            continue
        lines += 1
        row = labels[line]
        for m_idx in range(len(expected.index)):
            col_letter = get_column_letter(2 + m_idx)
            ref = f"'[{path.name}]MODEL'!{col_letter}{row}"
            if ref not in solution:
                failures.append(f"{line} month {m_idx + 1}: cell not evaluated (key {ref!r} missing)")
                continue
            raw = solution[ref].value[0][0]
            if raw is None or (isinstance(raw, float) and math.isnan(raw)):
                failures.append(f"{line} month {m_idx + 1}: workbook returned NaN/None")
                continue
            got = float(raw)
            want = float(expected[line].iloc[m_idx])
            denom = max(abs(want), 1.0)
            dev = abs(got - want) / denom
            max_dev = max(max_dev, dev)
            if not dev <= rel_tol:
                failures.append(
                    f"{line} month {m_idx + 1}: workbook {got!r} vs engine {want!r}"
                    f" (rel_dev={dev:.2e})"
                )

    return VerifyReport(
        passed=not failures and lines > 0,
        failures=failures,
        max_rel_deviation=max_dev,
        lines_checked=lines,
    )
