"""Canonical monthly translator: EntityConfig -> two-sheet live-formula workbook.

Assumptions sheet: named driver cells (editable).
Model sheet: every line a formula referencing those names.
Formula vocabulary: arithmetic, ^, SUM, MIN, MAX, IF only.
"""
from __future__ import annotations

from pathlib import Path
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pyfpa.config.schemas import EntityConfig
from pyfpa.excel.toolkit import (
    add_named_cell,
    add_named_row,
    days_format,
    fill_formula_row,
    freeze_header,
    money_format,
    percent_format,
)
from pyfpa.models.periods import month_index


class _SeasonRef(NamedTuple):
    name: str           # defined-name "seasonality_ch{i}"
    col_letters: list[str]   # 12 Assumptions column letters (B onward)
    row: int


class _ChannelRef(NamedTuple):
    name_rev: str       # "rev_annual_ch{i}"
    name_growth: str    # "growth_ch{i}"
    name_cogs: str      # "cogs_pct_ch{i}"
    seasonality: _SeasonRef


class _DebtRef(NamedTuple):
    name_open: str      # "debt_open_{k}"
    name_rate: str      # "debt_rate_{k}"
    name_prin: str      # "debt_prin_{k}" (term_loan only; "" for loc)
    kind: str


_ASSUMP_COL = 2   # column B for assumption values
_SEASON_COL = 2   # seasonality rows also start at column B


def _build_assumptions(
    wb: Workbook, cfg: EntityConfig
) -> tuple[list[_ChannelRef], list[str], list[_DebtRef]]:
    """Populate Assumptions sheet and return per-section refs."""
    ws = wb["Assumptions"]
    row = 2   # row 1 used as title

    ws.cell(row=1, column=1, value="Assumptions")

    def nc(name: str, val: float, fmt: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=name)
        add_named_cell(wb, ws, name=name, row=row, col=_ASSUMP_COL, value=val, number_format=fmt)
        row += 1

    def nr(name: str, vals: list[float]) -> _SeasonRef:
        nonlocal row
        ws.cell(row=row, column=1, value=name)
        add_named_row(wb, ws, name=name, row=row, start_col=_SEASON_COL, values=vals)
        ref = _SeasonRef(
            name=name,
            col_letters=[get_column_letter(_SEASON_COL + j) for j in range(12)],
            row=row,
        )
        row += 1
        return ref

    mfmt = money_format()
    pfmt = percent_format()
    dfmt = days_format()

    channel_refs: list[_ChannelRef] = []
    for i, ch in enumerate(cfg.channels, start=1):
        nc(f"rev_annual_ch{i}", ch.annual_revenue, mfmt)
        nc(f"growth_ch{i}", ch.growth_rate, pfmt)
        nc(f"cogs_pct_ch{i}", ch.cogs_pct, pfmt)
        sr = nr(f"seasonality_ch{i}", list(ch.seasonality))
        channel_refs.append(_ChannelRef(
            name_rev=f"rev_annual_ch{i}",
            name_growth=f"growth_ch{i}",
            name_cogs=f"cogs_pct_ch{i}",
            seasonality=sr,
        ))

    opex_names: list[str] = []
    for j, line in enumerate(cfg.opex, start=1):
        if line.kind == "fixed":
            nm = f"opex_amount_{j}"
            nc(nm, line.monthly_amount, mfmt)
        else:
            nm = f"opex_pct_{j}"
            nc(nm, line.pct_of_revenue, pfmt)
        opex_names.append(nm)

    for name, val, fmt in [
        ("dso_days", cfg.working_capital.dso_days, dfmt),
        ("dio_days", cfg.working_capital.dio_days, dfmt),
        ("dpo_days", cfg.working_capital.dpo_days, dfmt),
        ("tax_rate", cfg.tax_rate, pfmt),
        ("da_monthly", cfg.da_monthly, mfmt),
        ("capex_monthly", cfg.capex_monthly, mfmt),
        ("open_cash", cfg.opening_balances.cash, mfmt),
        ("open_ar", cfg.opening_balances.ar, mfmt),
        ("open_ap", cfg.opening_balances.ap, mfmt),
        ("open_inventory", cfg.opening_balances.inventory, mfmt),
        ("open_nol", cfg.opening_balances.nol, mfmt),
    ]:
        nc(name, val, fmt)

    debt_refs: list[_DebtRef] = []
    for k, inst in enumerate(cfg.debt, start=1):
        nc(f"debt_open_{k}", inst.opening_balance, mfmt)
        nc(f"debt_rate_{k}", inst.annual_rate, pfmt)
        nm_prin = ""
        if inst.kind == "term_loan":
            nm_prin = f"debt_prin_{k}"
            nc(nm_prin, inst.monthly_principal, mfmt)
        debt_refs.append(_DebtRef(
            name_open=f"debt_open_{k}",
            name_rate=f"debt_rate_{k}",
            name_prin=nm_prin,
            kind=inst.kind,
        ))

    return channel_refs, opex_names, debt_refs


_MODEL_START_COL = 2   # column B = first month


def _season_cell(sr: _SeasonRef, cal_month_0: int) -> str:
    """Absolute cell reference for a 0-based calendar month's seasonality weight."""
    return f"Assumptions!${sr.col_letters[cal_month_0]}${sr.row}"


def _build_model(
    wb: Workbook,
    cfg: EntityConfig,
    channel_refs: list[_ChannelRef],
    opex_names: list[str],
    debt_refs: list[_DebtRef],
) -> None:
    ws = wb["Model"]
    idx = month_index(cfg.start_month, cfg.horizon_months)
    n = cfg.horizon_months
    sc = _MODEL_START_COL
    mfmt = money_format()

    row_counter = [2]   # start at row 2; row 1 is headers

    def alloc() -> int:
        r = row_counter[0]
        row_counter[0] += 1
        return r

    def emit_fn(label: str, template, fmt: str = mfmt) -> int:
        r = alloc()
        fill_formula_row(ws, row=r, label=label, start_col=sc, n_cols=n,
                         template=template, number_format=fmt)
        return r

    def emit_cells(label: str, formulas: list[str], fmt: str = mfmt) -> int:
        """Write per-month cells directly (used when formula depends on own row)."""
        r = alloc()
        ws.cell(row=r, column=1, value=label)
        for m_idx, formula in enumerate(formulas):
            ws.cell(row=r, column=sc + m_idx, value=formula).number_format = fmt
        return r

    # Header row 1: month labels
    ws.cell(row=1, column=1, value="")
    for m_idx, period in enumerate(idx):
        ws.cell(row=1, column=sc + m_idx, value=str(period))

    # -- Per-channel revenue --
    # Bake (weight_cell_ref, year_exponent) per month per channel
    ch_rev_rows: list[int] = []
    for i, (ch, cref) in enumerate(zip(cfg.channels, channel_refs)):
        baked = [
            (_season_cell(cref.seasonality, period.month - 1), m_idx // 12)
            for m_idx, period in enumerate(idx)
        ]

        def make_rev(cref_=cref, baked_=baked):
            def t(m: int, col: str) -> str:
                wref, yexp = baked_[m - 1]
                return (
                    f"={cref_.name_rev}"
                    f"*({wref}/SUM({cref_.seasonality.name}))"
                    f"*(1+{cref_.name_growth})^{yexp}"
                )
            return t

        ch_rev_rows.append(emit_fn(f"revenue_ch{i + 1}", make_rev()))

    # -- Revenue total --
    r_rev = emit_fn(
        "revenue",
        lambda m, col: f"=SUM({col}{ch_rev_rows[0]}:{col}{ch_rev_rows[-1]})",
    )

    # -- COGS total (explicit product sum across channels) --
    r_cogs = emit_fn(
        "cogs",
        lambda m, col: "=" + "+".join(
            f"{col}{ch_rev_rows[i]}*{cref.name_cogs}"
            for i, cref in enumerate(channel_refs)
        ),
    )

    # -- Gross profit --
    r_gp = emit_fn("gross_profit", lambda m, col: f"={col}{r_rev}-{col}{r_cogs}")

    # -- Opex per line then total --
    opex_rows: list[int] = []
    for j, (line, nm) in enumerate(zip(cfg.opex, opex_names)):
        if line.kind == "fixed":
            opex_rows.append(emit_fn(f"opex_{j + 1}", lambda m, col, nm_=nm: f"={nm_}"))
        else:
            opex_rows.append(emit_fn(
                f"opex_{j + 1}",
                lambda m, col, nm_=nm, r=r_rev: f"={col}{r}*{nm_}",
            ))

    if opex_rows:
        r_opex = emit_fn(
            "opex",
            lambda m, col: f"=SUM({col}{opex_rows[0]}:{col}{opex_rows[-1]})",
        )
    else:
        r_opex = emit_fn("opex", lambda m, col: "=0")

    # -- EBITDA and D&A --
    r_ebitda = emit_fn("ebitda", lambda m, col: f"={col}{r_gp}-{col}{r_opex}")
    r_da = emit_fn("da", lambda m, col: "=da_monthly")

    # -- Per-instrument debt rows (balance, interest, principal) --
    debt_int_rows: list[int] = []
    debt_prin_rows: list[int] = []

    for k, (inst, dref) in enumerate(zip(cfg.debt, debt_refs)):
        # Balance (AFTER payment) -- references its own row, so allocate then fill directly
        bal_row = alloc()
        ws.cell(row=bal_row, column=1, value=f"debt_balance_{k + 1}")
        for m_idx in range(n):
            if inst.kind == "term_loan":
                if m_idx == 0:
                    formula = f"={dref.name_open}-MIN({dref.name_prin},{dref.name_open})"
                else:
                    pc = get_column_letter(sc + m_idx - 1)
                    formula = f"={pc}{bal_row}-MIN({dref.name_prin},{pc}{bal_row})"
            else:
                if m_idx == 0:
                    formula = f"={dref.name_open}"
                else:
                    pc = get_column_letter(sc + m_idx - 1)
                    formula = f"={pc}{bal_row}"
            ws.cell(row=bal_row, column=sc + m_idx, value=formula).number_format = mfmt

        # Interest (on PRE-payment balance = prior balance or opening)
        int_formulas = [
            f"={dref.name_open}*{dref.name_rate}/12" if m_idx == 0
            else f"={get_column_letter(sc + m_idx - 1)}{bal_row}*{dref.name_rate}/12"
            for m_idx in range(n)
        ]
        int_row = emit_cells(f"interest_{k + 1}", int_formulas)
        debt_int_rows.append(int_row)

        # Principal
        prin_formulas: list[str] = []
        for m_idx in range(n):
            if inst.kind != "term_loan":
                prin_formulas.append("=0")
            elif m_idx == 0:
                prin_formulas.append(f"=MIN({dref.name_prin},{dref.name_open})")
            else:
                pc = get_column_letter(sc + m_idx - 1)
                prin_formulas.append(f"=MIN({dref.name_prin},{pc}{bal_row})")
        prin_row = emit_cells(f"principal_{k + 1}", prin_formulas)
        debt_prin_rows.append(prin_row)

    # -- Interest and principal totals --
    if debt_int_rows:
        r_int = emit_fn(
            "interest",
            lambda m, col: "=" + "+".join(f"{col}{r}" for r in debt_int_rows),
        )
    else:
        r_int = emit_fn("interest", lambda m, col: "=0")

    if debt_prin_rows:
        r_prin = emit_fn(
            "principal",
            lambda m, col: "=" + "+".join(f"{col}{r}" for r in debt_prin_rows),
        )
    else:
        r_prin = emit_fn("principal", lambda m, col: "=0")

    # -- Pretax income (EBIT - interest; EBIT = EBITDA - D&A) --
    r_pretax = emit_fn(
        "pretax_income",
        lambda m, col: f"={col}{r_ebitda}-{col}{r_da}-{col}{r_int}",
    )

    # -- NOL: three rows that cross-reference each other and prior month --
    # Allocate all three rows first, then fill
    nol_open_row = alloc()
    nol_used_row = alloc()
    nol_close_row = alloc()
    ws.cell(row=nol_open_row, column=1, value="nol_opening")
    ws.cell(row=nol_used_row, column=1, value="nol_used")
    ws.cell(row=nol_close_row, column=1, value="nol_closing")

    for m_idx in range(n):
        cl = get_column_letter(sc + m_idx)
        # nol_opening
        nol_o = "=open_nol" if m_idx == 0 else f"={get_column_letter(sc + m_idx - 1)}{nol_close_row}"
        ws.cell(row=nol_open_row, column=sc + m_idx, value=nol_o).number_format = mfmt
        # nol_used
        ws.cell(
            row=nol_used_row, column=sc + m_idx,
            value=f"=MIN({cl}{nol_open_row},MAX(0,{cl}{r_pretax}))",
        ).number_format = mfmt
        # nol_closing
        ws.cell(
            row=nol_close_row, column=sc + m_idx,
            value=f"={cl}{nol_open_row}-{cl}{nol_used_row}",
        ).number_format = mfmt

    # -- Tax, net income --
    r_tax = emit_fn(
        "tax",
        lambda m, col: f"=(MAX(0,{col}{r_pretax})-{col}{nol_used_row})*tax_rate",
    )
    r_ni = emit_fn("net_income", lambda m, col: f"={col}{r_pretax}-{col}{r_tax}")

    # -- Working capital balances --
    r_ar = emit_fn("ar_balance", lambda m, col: f"={col}{r_rev}*dso_days/30")
    r_ap = emit_fn("ap_balance", lambda m, col: f"={col}{r_cogs}*dpo_days/30")
    r_inv = emit_fn("inv_balance", lambda m, col: f"={col}{r_cogs}*dio_days/30")

    # WC cash impact (first-month references opening balances)
    wc_formulas: list[str] = []
    for m_idx in range(n):
        cl = get_column_letter(sc + m_idx)
        if m_idx == 0:
            wc_formulas.append(
                f"=-({cl}{r_ar}-open_ar)"
                f"+({cl}{r_ap}-open_ap)"
                f"-({cl}{r_inv}-open_inventory)"
            )
        else:
            pc = get_column_letter(sc + m_idx - 1)
            wc_formulas.append(
                f"=-({cl}{r_ar}-{pc}{r_ar})"
                f"+({cl}{r_ap}-{pc}{r_ap})"
                f"-({cl}{r_inv}-{pc}{r_inv})"
            )
    r_wc = emit_cells("wc_cash_impact", wc_formulas)

    # -- Operating cash flow, capex, FCF, change in cash --
    r_ocf = emit_fn(
        "operating_cash_flow",
        lambda m, col: f"={col}{r_ni}+{col}{r_da}+{col}{r_wc}",
    )
    r_capex = emit_fn("capex", lambda m, col: "=capex_monthly")
    r_fcf = emit_fn("free_cash_flow", lambda m, col: f"={col}{r_ocf}-{col}{r_capex}")
    r_chg = emit_fn("change_in_cash", lambda m, col: f"={col}{r_fcf}-{col}{r_prin}")

    # Ending cash (cumulative, references own prior cell)
    end_row = alloc()
    ws.cell(row=end_row, column=1, value="ending_cash")
    for m_idx in range(n):
        cl = get_column_letter(sc + m_idx)
        if m_idx == 0:
            formula = f"=open_cash+{cl}{r_chg}"
        else:
            pc = get_column_letter(sc + m_idx - 1)
            formula = f"={pc}{end_row}+{cl}{r_chg}"
        ws.cell(row=end_row, column=sc + m_idx, value=formula).number_format = mfmt

    freeze_header(ws, first_data_cell="B2")


def model_to_excel(cfg: EntityConfig, path: str | Path) -> None:
    """Compile an EntityConfig into a two-sheet live-formula workbook at `path`."""
    wb = Workbook()
    ws_assump = wb.active
    ws_assump.title = "Assumptions"
    wb.create_sheet("Model")

    channel_refs, opex_names, debt_refs = _build_assumptions(wb, cfg)
    _build_model(wb, cfg, channel_refs, opex_names, debt_refs)

    wb.save(Path(path))
