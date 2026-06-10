from __future__ import annotations

from collections.abc import Callable, Sequence

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet


def money_format() -> str:
    return "#,##0"


def percent_format() -> str:
    return "0.0%"


def days_format() -> str:
    return "0.0"


def add_named_cell(
    wb: Workbook,
    ws: Worksheet,
    *,
    name: str,
    row: int,
    col: int,
    value: float,
    number_format: str | None = None,
) -> None:
    """Write a value and register a workbook-scoped defined name for the cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if number_format:
        cell.number_format = number_format
    ref = f"'{ws.title}'!${get_column_letter(col)}${row}"
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


def add_named_row(
    wb: Workbook,
    ws: Worksheet,
    *,
    name: str,
    row: int,
    start_col: int,
    values: Sequence[float],
    number_format: str | None = None,
) -> None:
    """Write a horizontal run of values and register a defined name for the range."""
    for offset, value in enumerate(values):
        cell = ws.cell(row=row, column=start_col + offset, value=value)
        if number_format:
            cell.number_format = number_format
    first = f"${get_column_letter(start_col)}${row}"
    last = f"${get_column_letter(start_col + len(values) - 1)}${row}"
    wb.defined_names[name] = DefinedName(
        name=name, attr_text=f"'{ws.title}'!{first}:{last}"
    )


def fill_formula_row(
    ws: Worksheet,
    *,
    row: int,
    label: str,
    start_col: int,
    n_cols: int,
    template: Callable[[int, str], str],
    number_format: str | None = None,
) -> None:
    """Label column A, then fill each month cell with template(month_index, column_letter).

    template receives the 1-based month index and that month's column letter and
    must return a full formula string starting with '='."""
    ws.cell(row=row, column=1, value=label)
    for m in range(1, n_cols + 1):
        col = get_column_letter(start_col + m - 1)
        cell = ws.cell(row=row, column=start_col + m - 1, value=template(m, col))
        if number_format:
            cell.number_format = number_format


def freeze_header(ws: Worksheet, *, first_data_cell: str = "B2") -> None:
    ws.freeze_panes = first_data_cell
