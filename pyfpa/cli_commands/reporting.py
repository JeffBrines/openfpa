from __future__ import annotations

import argparse

from openpyxl import load_workbook

from pyfpa.cli_helpers import _failure, _root, _success
from pyfpa.memory.workspace import workspace_path


def command_model_export(args: argparse.Namespace) -> int:
    from pyfpa.config.loader import load_config
    from pyfpa.excel.model_workbook import model_to_excel

    root = _root(args.path)
    workspace = workspace_path(root)
    if not workspace.is_dir():
        return _failure(
            "model-export",
            root,
            "workspace_not_initialized",
            "initialize the company workspace before exporting a model workbook",
        )
    try:
        cfg = load_config(args.config)
    except Exception as exc:
        return _failure("model-export", root, "invalid_config", str(exc))
    try:
        out_path = _root(args.out)
        model_to_excel(cfg, out_path)
        wb = load_workbook(out_path)
        sheet_names = wb.sheetnames
    except Exception as exc:
        return _failure("model-export", root, "export_failed", str(exc))
    return _success(
        "model-export",
        root,
        {
            "workbook_path": str(out_path),
            "sheets": sheet_names,
            "config": str(args.config),
            "entity": cfg.name,
        },
    )
