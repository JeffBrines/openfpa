# tests/test_excel_toolkit.py
from openpyxl import Workbook, load_workbook

from pyfpa.excel.toolkit import (
    add_named_cell, add_named_row, fill_formula_row, money_format, percent_format,
)


def test_add_named_cell_registers_defined_name(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    add_named_cell(wb, ws, name="tax_rate", row=2, col=2, value=0.21, number_format=percent_format())
    path = tmp_path / "t.xlsx"
    wb.save(path)
    back = load_workbook(path)
    assert "tax_rate" in back.defined_names
    assert back["Assumptions"].cell(row=2, column=2).value == 0.21


def test_add_named_row_registers_range(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    add_named_row(wb, ws, name="seasonality_ch1", row=3, start_col=2,
                  values=[1.0] * 12, number_format=None)
    path = tmp_path / "t.xlsx"
    wb.save(path)
    back = load_workbook(path)
    assert "seasonality_ch1" in back.defined_names
    assert back["Assumptions"].cell(row=3, column=13).value == 1.0


def test_fill_formula_row_writes_formula_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    # template receives 1-based month index and the column letter for that month
    fill_formula_row(
        ws, row=5, label="gross_profit", start_col=2, n_cols=3,
        template=lambda m, col: f"={col}3-{col}4",
    )
    assert ws.cell(row=5, column=1).value == "gross_profit"
    assert ws.cell(row=5, column=2).value == "=B3-B4"
    assert ws.cell(row=5, column=4).value == "=D3-D4"


def test_money_and_percent_formats_are_strings():
    assert isinstance(money_format(), str)
    assert isinstance(percent_format(), str)
