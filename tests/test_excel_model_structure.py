from openpyxl import load_workbook

from pyfpa.config.schemas import EntityConfig
from pyfpa.excel.model_workbook import model_to_excel


def _cfg():
    return EntityConfig.model_validate({
        "name": "T", "start_month": "2026-01", "horizon_months": 12, "tax_rate": 0.21,
        "channels": [
            {"name": "A", "annual_revenue": 1_200_000.0, "growth_rate": 0.10,
             "seasonality": [1.0] * 12, "cogs_pct": 0.5},
            {"name": "B", "annual_revenue": 600_000.0, "growth_rate": 0.0,
             "seasonality": [1, 1, 1, 1, 1, 1, 2, 2, 2, 1, 1, 1], "cogs_pct": 0.4},
        ],
        "opex": [
            {"name": "salaries", "kind": "fixed", "monthly_amount": 30_000.0},
            {"name": "marketing", "kind": "variable", "pct_of_revenue": 0.05},
        ],
        "debt": [
            {"name": "term", "kind": "term_loan", "opening_balance": 500_000.0,
             "annual_rate": 0.08, "monthly_principal": 10_000.0},
            {"name": "loc", "kind": "loc", "opening_balance": 100_000.0, "annual_rate": 0.10},
        ],
        "working_capital": {"dso_days": 45.0, "dpo_days": 30.0, "dio_days": 60.0},
        "opening_balances": {"cash": 50_000.0, "ar": 100_000.0, "ap": 60_000.0,
                              "inventory": 120_000.0, "nol": 40_000.0},
        "da_monthly": 2_000.0, "capex_monthly": 5_000.0,
    })


def test_workbook_has_named_assumptions_and_formula_cells(tmp_path):
    path = tmp_path / "model.xlsx"
    model_to_excel(_cfg(), path)
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Assumptions", "Model"}
    for name in ["rev_annual_ch1", "growth_ch1", "cogs_pct_ch1", "seasonality_ch1",
                 "rev_annual_ch2", "dso_days", "dio_days", "dpo_days", "tax_rate",
                 "da_monthly", "capex_monthly", "open_cash", "open_nol",
                 "debt_open_1", "debt_rate_1", "debt_prin_1", "debt_open_2",
                 "opex_amount_1", "opex_pct_2"]:
        assert name in wb.defined_names, f"missing defined name: {name}"
    model = wb["Model"]
    labels = [model.cell(row=r, column=1).value for r in range(2, model.max_row + 1)]
    # Use engine column names for the aggregate rows (per the plan's label adjustment note)
    for required in ["revenue", "gross_profit", "ebitda", "pretax_income",
                     "tax", "net_income", "wc_cash_impact", "operating_cash_flow",
                     "free_cash_flow", "change_in_cash", "ending_cash"]:
        assert required in labels, f"missing model row: {required}"
    # every data cell in a known row is a formula string, not a value
    row_idx = labels.index("ending_cash") + 2
    for col in range(2, 2 + 12):
        v = model.cell(row=row_idx, column=col).value
        assert isinstance(v, str) and v.startswith("="), f"not a formula: {v!r}"


def test_formula_vocabulary_is_restricted(tmp_path):
    import re
    path = tmp_path / "model.xlsx"
    model_to_excel(_cfg(), path)
    wb = load_workbook(path)
    model = wb["Model"]
    allowed = {"SUM", "MIN", "MAX", "IF"}
    for row in model.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                for fn in re.findall(r"([A-Z]{2,})\s*\(", cell.value):
                    assert fn in allowed, f"forbidden function {fn} in {cell.value}"
