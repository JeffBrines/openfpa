import pandas as pd
import pytest

pytest.importorskip("formulas")

from pyfpa.config.schemas import EntityConfig
from pyfpa.excel.model_workbook import model_to_excel
from pyfpa.excel.verify import verify_workbook
from pyfpa.models.cashflow import cashflow_from_config


def _simple_cfg():
    return EntityConfig.model_validate({
        "name": "T", "start_month": "2026-01", "horizon_months": 6, "tax_rate": 0.0,
        "channels": [{"name": "A", "annual_revenue": 120_000.0, "growth_rate": 0.0,
                      "seasonality": [1.0] * 12, "cogs_pct": 0.5}],
        "opex": [], "debt": [],
        "working_capital": {"dso_days": 0.0, "dpo_days": 0.0, "dio_days": 0.0},
        "opening_balances": {"cash": 0.0},
    })


def test_verify_passes_on_faithful_workbook(tmp_path):
    cfg = _simple_cfg()
    path = tmp_path / "m.xlsx"
    model_to_excel(cfg, path)
    report = verify_workbook(path, cashflow_from_config(cfg))
    assert report.passed, report.failures


def test_verify_fails_on_corrupted_formula(tmp_path):
    from openpyxl import load_workbook
    cfg = _simple_cfg()
    path = tmp_path / "m.xlsx"
    model_to_excel(cfg, path)
    wb = load_workbook(path)
    model = wb["Model"]
    labels = {model.cell(row=r, column=1).value: r for r in range(2, model.max_row + 1)}
    model.cell(row=labels["gross_profit"], column=3, value="=1234567")
    wb.save(path)
    report = verify_workbook(path, cashflow_from_config(cfg))
    assert not report.passed
    assert any("gross_profit" in f for f in report.failures)


def test_missing_formulas_dependency_message():
    import pyfpa.excel.verify as v
    import builtins
    real_import = builtins.__import__

    def fake_import(name, *args, **kwargs):
        if name == "formulas":
            raise ImportError("nope")
        return real_import(name, *args, **kwargs)

    builtins.__import__ = fake_import
    try:
        with pytest.raises(RuntimeError, match="pip install formulas"):
            v._load_formulas()
    finally:
        builtins.__import__ = real_import
