import json
import subprocess
import sys
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]


def run_cli(*args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, "-m", "pyfpa.cli", *args],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
    )


def output_json(result: subprocess.CompletedProcess[str]) -> dict:
    assert result.stdout, result.stderr
    return json.loads(result.stdout)


def test_model_export_writes_xlsx_with_both_sheets(tmp_path):
    from openpyxl import load_workbook

    assert run_cli("init", str(tmp_path)).returncode == 0
    config_path = ROOT / "examples/ridgeline/config.yaml"
    out_path = tmp_path / "model.xlsx"

    result = run_cli(
        "model-export",
        str(tmp_path),
        "--config",
        str(config_path),
        "--out",
        str(out_path),
    )

    assert result.returncode == 0
    payload = output_json(result)
    assert payload["ok"] is True
    assert payload["command"] == "model-export"
    data = payload["data"]
    assert set(data["sheets"]) == {"Assumptions", "Model"}
    assert out_path.exists()
    wb = load_workbook(out_path)
    assert set(wb.sheetnames) == {"Assumptions", "Model"}


def test_model_export_fails_without_workspace(tmp_path):
    config_path = ROOT / "examples/ridgeline/config.yaml"
    out_path = tmp_path / "model.xlsx"

    result = run_cli(
        "model-export",
        str(tmp_path),
        "--config",
        str(config_path),
        "--out",
        str(out_path),
    )

    assert result.returncode == 1
    payload = output_json(result)
    assert payload["ok"] is False
    assert payload["error"]["type"] == "workspace_not_initialized"


def test_model_export_fails_on_invalid_config(tmp_path):
    assert run_cli("init", str(tmp_path)).returncode == 0
    bad_config = tmp_path / "bad.yaml"
    bad_config.write_text("not: valid: entity: config\n")
    out_path = tmp_path / "model.xlsx"

    result = run_cli(
        "model-export",
        str(tmp_path),
        "--config",
        str(bad_config),
        "--out",
        str(out_path),
    )

    assert result.returncode == 1
    payload = output_json(result)
    assert payload["ok"] is False
